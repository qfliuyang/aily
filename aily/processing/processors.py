"""Universal content processors.

Each processor handles a specific content type and extracts clean text.
"""

from __future__ import annotations

import io
import logging
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from aily.llm.client import LLMClient

logger = logging.getLogger(__name__)


def _get_title_from_filename(filename: str | None) -> str | None:
    """Extract title from filename, returning None if filename is empty."""
    return Path(filename).stem if filename else None


@dataclass
class ExtractedContent:
    """Unified output format for all processors."""

    text: str
    title: str | None = None
    source_type: str = "unknown"
    metadata: dict = None  # page count, OCR confidence, etc.

    def __post_init__(self):
        if self.metadata is None:
            self.metadata = {}


class ContentProcessor(ABC):
    """Base class for all content processors."""

    SUPPORTED_TYPES: list[str] = []

    @abstractmethod
    async def process(self, data: bytes, filename: str | None = None) -> ExtractedContent:
        """Process content and extract text.

        Args:
            data: Raw file bytes
            filename: Original filename (optional)

        Returns:
            ExtractedContent with text and metadata
        """
        pass

    def can_process(self, mime_type: str) -> bool:
        """Check if this processor handles the given MIME type."""
        # Support wildcards like "image/*"
        for supported in self.SUPPORTED_TYPES:
            if supported.endswith("/*"):
                prefix = supported[:-1]
                if mime_type.startswith(prefix):
                    return True
            elif supported == mime_type:
                return True
        return False


class PDFProcessor(ContentProcessor):
    """Extract text from PDF documents."""

    SUPPORTED_TYPES = ["application/pdf"]

    def __init__(self, extract_tables: bool = False) -> None:
        self.extract_tables = extract_tables
        self._pdfplumber_available = None

    async def process(self, data: bytes, filename: str | None = None) -> ExtractedContent:
        """Extract text from PDF using pdfplumber."""
        if not self._is_pdfplumber_available():
            logger.error("pdfplumber not installed. Run: pip install pdfplumber")
            return ExtractedContent(
                text="[PDF processing unavailable - pdfplumber not installed]",
                source_type="pdf",
            )

        import pdfplumber

        text_parts = []
        page_count = 0

        try:
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                page_count = len(pdf.pages)
                for i, page in enumerate(pdf.pages):
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(f"\n--- Page {i + 1} ---\n")
                        text_parts.append(page_text)

            full_text = "\n".join(text_parts)

            # Clean up excessive whitespace
            full_text = re.sub(r"\n{3,}", "\n\n", full_text)

            return ExtractedContent(
                text=full_text.strip(),
                title=_get_title_from_filename(filename),
                source_type="pdf",
                metadata={
                    "page_count": page_count,
                    "filename": filename,
                },
            )

        except Exception as e:
            logger.exception("PDF extraction failed")
            return ExtractedContent(
                text=f"[PDF extraction failed: {e}]",
                source_type="pdf",
            )

    def _is_pdfplumber_available(self) -> bool:
        if self._pdfplumber_available is None:
            try:
                import pdfplumber

                self._pdfplumber_available = True
            except ImportError:
                self._pdfplumber_available = False
        return self._pdfplumber_available


class ImageProcessor(ContentProcessor):
    """Extract text from images using OCR."""

    SUPPORTED_TYPES = ["image/png", "image/jpeg", "image/gif", "image/webp", "image/bmp"]

    def __init__(self, languages: list[str] | None = None) -> None:
        # Default includes English and Chinese
        self.languages = languages or ["en", "ch_sim"]
        self._easyocr_available = None
        self._reader = None

    async def process(self, data: bytes, filename: str | None = None) -> ExtractedContent:
        """Extract text from image using EasyOCR."""
        if not self._is_easyocr_available():
            logger.error("EasyOCR not installed. Run: pip install easyocr")
            return ExtractedContent(
                text="[Image OCR unavailable - easyocr not installed]",
                source_type="image",
            )

        try:
            import easyocr
            from PIL import Image

            # Lazy init reader (expensive to load models)
            if self._reader is None:
                logger.info("Initializing EasyOCR (first use)...")
                self._reader = easyocr.Reader(self.languages, gpu=False)

            # Load image from bytes
            image = Image.open(io.BytesIO(data))

            # Convert to format EasyOCR expects
            if image.mode != "RGB":
                image = image.convert("RGB")

            # Run OCR in thread pool to avoid blocking event loop
            results = await asyncio.to_thread(self._reader.readtext, data)

            # Extract text (results are [(bbox, text, confidence), ...])
            text_parts = [result[1] for result in results]
            full_text = "\n".join(text_parts)

            # Calculate average confidence
            confidences = [result[2] for result in results]
            avg_confidence = sum(confidences) / len(confidences) if confidences else 0

            return ExtractedContent(
                text=full_text.strip(),
                title=_get_title_from_filename(filename),
                source_type="image",
                metadata={
                    "ocr_confidence": round(avg_confidence, 3),
                    "text_blocks": len(text_parts),
                    "filename": filename,
                    "image_size": image.size,
                },
            )

        except Exception as e:
            logger.exception("Image OCR failed")
            return ExtractedContent(
                text=f"[Image OCR failed: {e}]",
                source_type="image",
            )

    def _is_easyocr_available(self) -> bool:
        if self._easyocr_available is None:
            try:
                import easyocr

                self._easyocr_available = True
            except ImportError:
                self._easyocr_available = False
        return self._easyocr_available


class MarkdownProcessor(ContentProcessor):
    """Process markdown files - extract text and metadata."""

    SUPPORTED_TYPES = ["text/markdown", "text/x-markdown"]

    def __init__(self) -> None:
        self._frontmatter_pattern = re.compile(
            r"^---\s*\n(.*?)\n---\s*\n", re.DOTALL
        )

    async def process(self, data: bytes, filename: str | None = None) -> ExtractedContent:
        """Parse markdown, extracting frontmatter and content."""
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            text = data.decode("utf-8", errors="ignore")

        # Extract frontmatter
        frontmatter = {}
        content = text

        match = self._frontmatter_pattern.match(text)
        if match:
            frontmatter_text = match.group(1)
            content = text[match.end() :]

            # Simple YAML-like parsing (key: value)
            for line in frontmatter_text.strip().split("\n"):
                if ":" in line:
                    key, value = line.split(":", 1)
                    frontmatter[key.strip()] = value.strip().strip('"').strip("'")

        # Extract title from frontmatter or first heading
        title = frontmatter.get("title")
        if not title:
            # Try first # heading
            heading_match = re.search(r"^#\s+(.+)$", content, re.MULTILINE)
            if heading_match:
                title = heading_match.group(1).strip()
            else:
                title = _get_title_from_filename(filename)

        return ExtractedContent(
            text=content.strip(),
            title=title,
            source_type="markdown",
            metadata={
                "frontmatter": frontmatter,
                "filename": filename,
            },
        )


class DocxProcessor(ContentProcessor):
    """Extract text from Word documents (.docx)."""

    SUPPORTED_TYPES = [
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ]

    async def process(self, data: bytes, filename: str | None = None) -> ExtractedContent:
        """Extract text from DOCX using python-docx."""
        try:
            from docx import Document
        except ImportError:
            logger.error("python-docx not installed. Run: pip install python-docx")
            return ExtractedContent(
                text="[DOCX processing unavailable - python-docx not installed]",
                source_type="docx",
            )

        try:
            doc = Document(io.BytesIO(data))

            text_parts = []
            for para in doc.paragraphs:
                if para.text.strip():
                    text_parts.append(para.text)

            full_text = "\n\n".join(text_parts)

            return ExtractedContent(
                text=full_text.strip(),
                title=_get_title_from_filename(filename),
                source_type="docx",
                metadata={
                    "paragraph_count": len(text_parts),
                    "filename": filename,
                },
            )

        except Exception as e:
            logger.exception("DOCX extraction failed")
            return ExtractedContent(
                text=f"[DOCX extraction failed: {e}]",
                source_type="docx",
            )


class TextProcessor(ContentProcessor):
    """Plain text pass-through processor."""

    SUPPORTED_TYPES = ["text/plain", "text/*"]

    async def process(self, data: bytes, filename: str | None = None) -> ExtractedContent:
        """Process plain text - just decode and return."""
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            text = data.decode("utf-8", errors="ignore")

        return ExtractedContent(
            text=text.strip(),
            title=_get_title_from_filename(filename),
            source_type="text",
            metadata={"filename": filename},
        )


class WebProcessor(ContentProcessor):
    """Process web pages (HTML) - delegates to browser fetcher."""

    SUPPORTED_TYPES = ["text/html", "application/xhtml+xml"]

    def __init__(self, browser_manager=None) -> None:
        self.browser_manager = browser_manager

    async def process(self, data: bytes, filename: str | None = None) -> ExtractedContent:
        """Process HTML content.

        If browser_manager is available, use it for JS-rendered pages.
        Otherwise do basic HTML parsing.
        """
        # For now, basic HTML text extraction
        # In production, this would use the browser manager for full JS support
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            text = data.decode("utf-8", errors="ignore")

        # Very basic HTML stripping
        # Remove script/style tags
        text = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.DOTALL)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL)

        # Extract title
        title_match = re.search(r"<title[^>]*>(.*?)</title>", text, re.DOTALL)
        title = title_match.group(1).strip() if title_match else None

        # Strip remaining HTML tags
        text = re.sub(r"<[^>]+>", " ", text)

        # Clean up whitespace
        text = re.sub(r"\s+", " ", text)

        return ExtractedContent(
            text=text.strip(),
            title=title,
            source_type="web",
            metadata={"filename": filename},
        )


class CSVProcessor(ContentProcessor):
    """Extract text from CSV files."""

    SUPPORTED_TYPES = ["text/csv"]

    async def process(self, data: bytes, filename: str | None = None) -> ExtractedContent:
        """Extract text from CSV using the csv module."""
        try:
            import csv
            import io

            # Decode the CSV data
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                text = data.decode("utf-8", errors="ignore")

            # Parse CSV
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)

            if not rows:
                return ExtractedContent(
                    text="[Empty CSV file]",
                    title=_get_title_from_filename(filename),
                    source_type="csv",
                    metadata={"filename": filename, "row_count": 0},
                )

            # Format as markdown table
            lines = []
            for i, row in enumerate(rows):
                # Escape pipe characters in cells
                escaped = [cell.replace("|", "\\|") for cell in row]
                lines.append("| " + " | ".join(escaped) + " |")
                # Add separator after header
                if i == 0:
                    lines.append("|" + "|".join(" --- " for _ in row) + "|")

            full_text = "\n".join(lines)

            return ExtractedContent(
                text=full_text,
                title=_get_title_from_filename(filename),
                source_type="csv",
                metadata={
                    "filename": filename,
                    "row_count": len(rows),
                    "column_count": len(rows[0]) if rows else 0,
                },
            )

        except Exception as e:
            logger.exception("CSV extraction failed")
            return ExtractedContent(
                text=f"[CSV extraction failed: {e}]",
                source_type="csv",
            )


class XLSXProcessor(ContentProcessor):
    """Extract text from Excel files (.xlsx)."""

    SUPPORTED_TYPES = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ]

    async def process(self, data: bytes, filename: str | None = None) -> ExtractedContent:
        """Extract text from XLSX using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return ExtractedContent(
                text="[XLSX processing unavailable - openpyxl not installed]",
                source_type="xlsx",
            )

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

            all_sheets = []
            total_rows = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = []

                for row in sheet.iter_rows(values_only=True):
                    # Convert None to empty string, ensure all values are strings
                    cell_values = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell_values):  # Skip completely empty rows
                        # Escape pipe characters
                        escaped = [cell.replace("|", "\\|") for cell in cell_values]
                        rows.append("| " + " | ".join(escaped) + " |")
                        total_rows += 1

                if rows:
                    # Add separator after header (first row)
                    if len(rows) > 0:
                        # Calculate column count from first row
                        col_count = len(rows[0].split("|")) - 2  # Remove empty splits
                        separator = "|" + "|".join(" --- " for _ in range(col_count)) + "|"
                        rows.insert(1, separator)

                    all_sheets.append(f"## Sheet: {sheet_name}\n")
                    all_sheets.append("\n".join(rows))
                    all_sheets.append("\n")

            full_text = "\n".join(all_sheets)

            return ExtractedContent(
                text=full_text.strip(),
                title=_get_title_from_filename(filename),
                source_type="xlsx",
                metadata={
                    "filename": filename,
                    "sheet_count": len(workbook.sheetnames),
                    "total_rows": total_rows,
                },
            )

        except Exception as e:
            logger.exception("XLSX extraction failed")
            return ExtractedContent(
                text=f"[XLSX extraction failed: {e}]",
                source_type="xlsx",
            )
