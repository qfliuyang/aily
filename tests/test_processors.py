import pytest
from unittest.mock import patch, MagicMock, AsyncMock
import io

from aily.processing.processors import (
    ExtractedContent,
    PDFProcessor,
    ImageProcessor,
    MarkdownProcessor,
    DocxProcessor,
    TextProcessor,
    WebProcessor,
    CSVProcessor,
    XLSXProcessor,
    _get_title_from_filename,
)


class TestGetTitleFromFilename:
    def test_returns_stem_for_valid_filename(self):
        assert _get_title_from_filename("document.pdf") == "document"
        assert _get_title_from_filename("/path/to/file.txt") == "file"

    def test_returns_none_for_empty_filename(self):
        assert _get_title_from_filename(None) is None
        assert _get_title_from_filename("") is None


class TestExtractedContent:
    def test_post_init_creates_empty_metadata(self):
        content = ExtractedContent(text="hello")
        assert content.metadata == {}

    def test_preserves_provided_metadata(self):
        content = ExtractedContent(text="hello", metadata={"key": "value"})
        assert content.metadata == {"key": "value"}


class TestPDFProcessor:
    def test_supported_types(self):
        processor = PDFProcessor()
        assert processor.can_process("application/pdf")
        assert not processor.can_process("text/plain")

    @pytest.mark.asyncio
    async def test_returns_error_when_pdfplumber_not_installed(self):
        processor = PDFProcessor()
        processor._pdfplumber_available = False
        result = await processor.process(b"fake pdf data", "test.pdf")
        assert "pdfplumber not installed" in result.text
        assert result.source_type == "pdf"

    @pytest.mark.asyncio
    @pytest.mark.skipif(
        __import__("importlib").util.find_spec("pdfplumber") is None,
        reason="pdfplumber not installed"
    )
    async def test_extracts_text_from_pdf(self):
        import pdfplumber
        processor = PDFProcessor()
        processor._pdfplumber_available = True

        mock_page = MagicMock()
        mock_page.extract_text.return_value = "Page 1 content"
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]

        with patch("pdfplumber.open", return_value=mock_pdf):
            result = await processor.process(b"fake pdf data", "test.pdf")
            assert "Page 1 content" in result.text
            assert result.title == "test"
            assert result.metadata["page_count"] == 1


class TestMarkdownProcessor:
    def test_supported_types(self):
        processor = MarkdownProcessor()
        assert processor.can_process("text/markdown")
        assert processor.can_process("text/x-markdown")
        assert not processor.can_process("text/plain")

    @pytest.mark.asyncio
    async def test_extracts_frontmatter_and_content(self):
        processor = MarkdownProcessor()
        md_data = b"""---
title: My Document
author: Test
---
# Heading

Some content.
"""
        result = await processor.process(md_data, "doc.md")
        assert result.title == "My Document"
        assert "Heading" in result.text
        assert result.metadata["frontmatter"]["title"] == "My Document"

    @pytest.mark.asyncio
    async def test_extracts_title_from_first_heading(self):
        processor = MarkdownProcessor()
        md_data = b"# First Heading\n\nContent here."
        result = await processor.process(md_data, "doc.md")
        assert result.title == "First Heading"

    @pytest.mark.asyncio
    async def test_falls_back_to_filename_for_title(self):
        processor = MarkdownProcessor()
        md_data = b"Just content, no heading."
        result = await processor.process(md_data, "mydoc.md")
        assert result.title == "mydoc"


class TestDocxProcessor:
    def test_supported_types(self):
        processor = DocxProcessor()
        assert processor.can_process(
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
        assert not processor.can_process("text/plain")

    @pytest.mark.asyncio
    async def test_returns_error_when_python_docx_not_installed(self):
        processor = DocxProcessor()
        with patch("builtins.__import__", side_effect=ImportError("No module named docx")):
            result = await processor.process(b"fake docx data", "test.docx")
            assert "python-docx not installed" in result.text


class TestTextProcessor:
    def test_supported_types(self):
        processor = TextProcessor()
        assert processor.can_process("text/plain")
        assert processor.can_process("text/csv")  # wildcard match
        assert not processor.can_process("application/pdf")

    @pytest.mark.asyncio
    async def test_decodes_utf8(self):
        processor = TextProcessor()
        result = await processor.process(b"Hello, World!", "test.txt")
        assert result.text == "Hello, World!"
        assert result.title == "test"
        assert result.source_type == "text"

    @pytest.mark.asyncio
    async def test_handles_invalid_utf8(self):
        processor = TextProcessor()
        result = await processor.process(b"\xff\xfe invalid utf8", "test.txt")
        assert "invalid utf8" in result.text  # Invalid bytes removed, rest preserved


class TestWebProcessor:
    def test_supported_types(self):
        processor = WebProcessor()
        assert processor.can_process("text/html")
        assert processor.can_process("application/xhtml+xml")
        assert not processor.can_process("text/plain")

    @pytest.mark.asyncio
    async def test_strips_html_tags(self):
        processor = WebProcessor()
        html = b"<html><head><title>Test Page</title></head><body><p>Hello World</p></body></html>"
        result = await processor.process(html, "page.html")
        assert result.title == "Test Page"
        assert "Hello World" in result.text
        assert "<html>" not in result.text

    @pytest.mark.asyncio
    async def test_removes_script_and_style(self):
        processor = WebProcessor()
        html = b"""<html>
<head><style>body { color: red; }</style></head>
<body>
<script>alert('hi');</script>
<p>Visible content</p>
</body>
</html>"""
        result = await processor.process(html, "page.html")
        assert "body { color: red; }" not in result.text
        assert "alert('hi')" not in result.text
        assert "Visible content" in result.text

    def test_is_js_required_domain(self):
        """Test detection of JS-required domains."""
        processor = WebProcessor()

        # Known JS-required domains
        assert processor._is_js_required_domain("https://monica.im/share/chat")
        assert processor._is_js_required_domain("https://chatgpt.com/c/123")
        assert processor._is_js_required_domain("https://claude.ai/chat/456")

        # Regular domains
        assert not processor._is_js_required_domain("https://example.com/page")
        assert not processor._is_js_required_domain("https://news.ycombinator.com")
        assert not processor._is_js_required_domain(None)

    def test_is_low_quality(self):
        """Test detection of low-quality content."""
        from aily.processing.processors import ExtractedContent

        processor = WebProcessor()

        # Very short content
        short = ExtractedContent(text="Short", source_type="web")
        assert processor._is_low_quality(short)

        # Generic Monica title
        monica = ExtractedContent(
            text="Monica - Your GPT AI Assistant Chrome Extension",
            title="Monica - Your GPT AI Assistant Chrome Extension",
            source_type="web"
        )
        assert processor._is_low_quality(monica)

        # Good content
        good = ExtractedContent(
            text="This is a long article with lots of content about AI and machine learning. " * 20,
            title="Understanding Neural Networks",
            source_type="web"
        )
        assert not processor._is_low_quality(good)

        # Empty content
        empty = ExtractedContent(text="", source_type="web")
        assert processor._is_low_quality(empty)

    @pytest.mark.asyncio
    async def test_detects_monica_share_links(self):
        """Test that Monica share links are detected as low quality."""
        processor = WebProcessor()

        # Simulate Monica's generic response
        html = b"""<html>
<head><title>Monica - Your GPT AI Assistant Chrome Extension</title></head>
<body>
<p>Monica is a GPT-4 powered AI assistant.</p>
</body>
</html>"""

        result = await processor.process(
            html,
            filename="chat.html",
            url="https://monica.im/share/chat?shareId=abc123"
        )

        # Should detect generic title
        assert result.title and "monica" in result.title.lower()
        # Content should be flagged as low quality
        assert processor._is_low_quality(result)


class TestCSVProcessor:
    def test_supported_types(self):
        processor = CSVProcessor()
        assert processor.can_process("text/csv")
        assert not processor.can_process("text/plain")

    @pytest.mark.asyncio
    async def test_converts_csv_to_markdown_table(self):
        processor = CSVProcessor()
        csv_data = b"name,age,city\nAlice,30,NYC\nBob,25,LA"
        result = await processor.process(csv_data, "data.csv")
        assert "| name | age | city |" in result.text
        assert "| Alice | 30 | NYC |" in result.text
        assert "| Bob | 25 | LA |" in result.text
        assert result.title == "data"
        assert result.metadata["row_count"] == 3
        assert result.metadata["column_count"] == 3

    @pytest.mark.asyncio
    async def test_handles_empty_csv(self):
        processor = CSVProcessor()
        result = await processor.process(b"", "empty.csv")
        assert "[Empty CSV file]" in result.text

    @pytest.mark.asyncio
    async def test_escapes_pipe_characters(self):
        processor = CSVProcessor()
        csv_data = b"col1,col2\nval|ue1,value2"
        result = await processor.process(csv_data, "data.csv")
        assert "val\\|ue1" in result.text  # Pipe should be escaped


class TestXLSXProcessor:
    def test_supported_types(self):
        processor = XLSXProcessor()
        assert processor.can_process(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert not processor.can_process("text/plain")

    @pytest.mark.asyncio
    async def test_returns_error_when_openpyxl_not_installed(self):
        processor = XLSXProcessor()
        with patch("builtins.__import__", side_effect=ImportError("No module named openpyxl")):
            result = await processor.process(b"fake xlsx data", "test.xlsx")
            assert "openpyxl not installed" in result.text

    @pytest.mark.asyncio
    @pytest.mark.skipif(
        __import__("importlib").util.find_spec("openpyxl") is None,
        reason="openpyxl not installed"
    )
    async def test_extracts_sheets_as_markdown(self):
        import openpyxl
        processor = XLSXProcessor()

        mock_sheet = MagicMock()
        mock_sheet.iter_rows.return_value = [
            ["Name", "Age"],
            ["Alice", 30],
            ["Bob", 25],
        ]

        mock_workbook = MagicMock()
        mock_workbook.sheetnames = ["Sheet1"]
        mock_workbook.__getitem__ = MagicMock(return_value=mock_sheet)

        with patch("openpyxl.load_workbook", return_value=mock_workbook):
            result = await processor.process(b"fake xlsx data", "data.xlsx")
            assert "## Sheet: Sheet1" in result.text
            assert "| Name | Age |" in result.text
            assert "| Alice | 30 |" in result.text
            assert result.metadata["sheet_count"] == 1


class TestImageProcessor:
    def test_supported_types(self):
        processor = ImageProcessor()
        assert processor.can_process("image/png")
        assert processor.can_process("image/jpeg")
        assert processor.can_process("image/webp")
        assert not processor.can_process("text/plain")

    @pytest.mark.asyncio
    async def test_returns_error_when_easyocr_not_installed(self):
        processor = ImageProcessor()
        processor._easyocr_available = False
        result = await processor.process(b"fake image data", "test.png")
        assert "easyocr not installed" in result.text
